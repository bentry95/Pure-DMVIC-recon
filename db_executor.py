import csv
import time
import os
import datetime
from typing import Callable, List, Tuple, Optional, Dict, Any

try:
    import pyodbc
except ImportError:
    pyodbc = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

COMMON_SQL_DRIVERS = [
    "ODBC Driver 17 for SQL Server",
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 13 for SQL Server",
    "ODBC Driver 11 for SQL Server",
    "SQL Server Native Client 11.0",
    "SQL Server"
]

DEFAULT_QUERY = """
SELECT a.insurance_ref AS PolicyNumber,
       a.insurance_holder_name AS Insured,
       REPLACE(CONVERT(varchar(11), a.cover_start_date, 103), '-', '/') AS StartDate,
       REPLACE(CONVERT(varchar(11), a.expiry_date, 103), '-', '/') AS ExpiryDate,
       e.description AS Class,
       CASE WHEN a.agent_shortname IS NULL THEN 'DIRECT' ELSE a.agent_shortname END AS Agent,
       a.document_ref AS DocumentNo,
       YEAR(a.cover_start_date) AS Year,
       REPLACE(CONVERT(varchar(11), a.transaction_date, 103), '-', '/') AS DocumentDate,
       CASE WHEN a.source_id = 11 THEN 'International Business' ELSE ISNULL(k.name, 'Direct') END AS Channel,
       CASE WHEN h.description = 'MTA' THEN 'Mid Term Adjustment' ELSE h.description END AS TransType,
       CASE b.stats_detail_type WHEN 'GRS' THEN b.sum_insured_home ELSE 0 END AS SumInsured,
       CASE b.stats_detail_type WHEN 'GRS' THEN b.this_premium_home ELSE 0 END AS GrossPrem,
       CASE b.stats_detail_type WHEN 'TTY' THEN b.this_premium_home ELSE 0 END AS TreatyPrem,
       CASE b.stats_detail_type WHEN 'FAC' THEN b.this_premium_home ELSE 0 END AS FacPrem,
       CASE b.stats_detail_type WHEN 'NET' THEN b.this_premium_home * -1 ELSE 0 END AS NetPremium,
       g.code AS Branch,
       a.posting_period_number AS Period,
       a.document_ref AS DebitNote,
       a.cover_start_date AS PureStartDate,
       a.expiry_date AS PureEndDate
FROM (Party d
RIGHT OUTER JOIN ((Period c 
INNER JOIN ((Stats_Detail b 
INNER JOIN Stats_Folder a        ON b.stats_folder_cnt = a.stats_folder_cnt
INNER JOIN transaction_type h    ON h.transaction_type_id = a.transaction_type_id) 
INNER JOIN Class_Of_Business e   ON b.class_of_business_id = e.class_of_business_id)    ON c.period_id = a.posting_period_number) 
INNER JOIN Insurance_File f        ON a.insurance_file_cnt = f.insurance_file_cnt)    ON d.party_cnt = f.lead_agent_cnt) 
INNER JOIN Source g                ON f.source_id = g.source_id
 LEFT JOIN Party i                ON i.party_cnt = a.agent_cnt
 LEFT JOIN Party_Agent j          ON j.party_cnt = i.party_cnt
 LEFT JOIN Party k                ON k.party_cnt = j.linked_account_group
WHERE a.posting_period_year >= YEAR(?) AND a.posting_period_year <= YEAR(?)
  AND a.document_ref LIKE 'S%'
GROUP BY PolicyNumber, Insured, StartDate, ExpiryDate, Class, Agent, DocumentNo, Year, DocumentDate, Branch, Channel, TransType, SumInsured, GrossPrem, TreatyPrem, FacPrem, NetPremium, DebitNote, PureStartDate, PureEndDate, Period
ORDER BY Period, DocumentDate, DocumentNo;
"""

def get_available_drivers() -> List[str]:
    """Return list of installed ODBC drivers matching SQL Server."""
    if pyodbc is None:
        return COMMON_SQL_DRIVERS
    installed = pyodbc.drivers()
    sql_drivers = [d for d in installed if "SQL" in d.upper() or "SQLSERVER" in d.upper()]
    return sql_drivers if sql_drivers else COMMON_SQL_DRIVERS

def build_connection_string(server: str, database: str, username: str, password: str, auth_type: str, driver: str) -> str:
    """Build connection string."""
    server = server.strip()
    database = database.strip() or "SIRIUS_MW"
    username = username.strip()
    driver = driver.strip()

    if auth_type == "Windows Authentication":
        conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};Trusted_Connection=yes;TrustServerCertificate=yes;"
    else:
        conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password};TrustServerCertificate=yes;"
    
    return conn_str

def test_connection(server: str, database: str, username: str, password: str, auth_type: str, driver: str) -> Tuple[bool, str]:
    """Test system connection."""
    if pyodbc is None:
        return False, "System component missing."
    
    conn_str = build_connection_string(server, database, username, password, auth_type, driver)
    try:
        conn = pyodbc.connect(conn_str, timeout=10)
        conn.close()
        return True, "System connection ready."
    except Exception as e:
        return False, f"System Connection Error: {str(e)}"

def run_filtered_query(
    server: str,
    database: str,
    username: str,
    password: str,
    auth_type: str,
    driver: str,
    start_date: str,
    end_date: str,
    log_callback: Callable[[str], None]
) -> Tuple[bool, List[str], List[Tuple[Any, ...]], str]:
    """
    Executes internal report lookup for the specified date range (debit_credit = 'D').
    """
    if pyodbc is None:
        log_callback("ERROR: Required system component missing.")
        return False, [], [], "component missing"

    log_callback(f"Retrieving system records for date range: {start_date} to {end_date} (debit_credit = 'D')...")
    conn_str = build_connection_string(server, database, username, password, auth_type, driver)
    
    start_time = time.time()
    try:
        conn = pyodbc.connect(conn_str, timeout=25)
        cursor = conn.cursor()
        cursor.execute(f"USE [{database}];")
        cursor.execute(DEFAULT_QUERY, (start_date, end_date))
        
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        conn.close()
        
        elapsed = time.time() - start_time
        log_callback(f"System lookup completed in {elapsed:.2f} seconds. Total matching records: {len(rows)}")
        return True, columns, rows, f"SUCCESS: Retrieved {len(rows)} record(s)."

    except Exception as query_err:
        err_msg = f"System Processing Error: {query_err}"
        log_callback(f"ERROR: {err_msg}")
        return False, [], [], err_msg

def clean_reg_no(reg_val: Any) -> str:
    """Helper to clean registration number string for matching."""
    if reg_val is None:
        return ""
    return str(reg_val).replace(" ", "").strip().upper()

def process_excel_pure_dmvic_recon(
    excel_path: str,
    output_excel_path: str,
    start_date: str,
    end_date: str,
    server: str,
    database: str,
    username: str,
    password: str,
    auth_type: str,
    driver: str,
    log_callback: Callable[[str], None]
) -> Tuple[bool, Any, int, int, str]:
    """
    PURE-DMVIC RECON (v1.0):
    1. Header 'DEBITNOTE' placed at cell T4 (Column T, Row 4).
    2. Data rows start at Row 5. Matches Registration Number for date range start_date to end_date (debit_credit = 'D').
    3. Writes corresponding DebitNote into cell T{r} for each row.
    """
    if openpyxl is None:
        log_callback("ERROR: Required component 'openpyxl' missing.")
        return False, None, 0, 0, "openpyxl missing"

    log_callback(f"Opening DMVIC report file: {excel_path}")
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        err_msg = f"Failed to load DMVIC report file: {e}"
        log_callback(f"ERROR: {err_msg}")
        return False, None, 0, 0, err_msg

    # Find Registration Number column index (searching rows 1-4 across all columns)
    reg_col_idx = None
    header_row_found = 4
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_clean = str(val).strip().lower()
                if ("reg" in val_clean and ("no" in val_clean or "number" in val_clean or "registration" in val_clean)) or ("registration" in val_clean):
                    reg_col_idx = c
                    header_row_found = r
                    log_callback(f"Found Registration Number column at Column {openpyxl.utils.get_column_letter(c)} (Row {r}: '{val}')")
                    break
        if reg_col_idx:
            break
    if reg_col_idx is None:
        reg_col_idx = 1
        log_callback("Registration Number header default to Column A.")

    # Attempt to locate a DMVIC date column (e.g., 'Document Date' or similar)
    date_col_idx = None
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_clean = str(val).strip().lower()
                if "date" in val_clean:
                    date_col_idx = c
                    log_callback(f"Found DMVIC date column at Column {openpyxl.utils.get_column_letter(c)} (Row {r}: '{val}')")
                    break
        if date_col_idx:
            break
    if date_col_idx is None:
        log_callback("No explicit DMVIC date column found; will leave DMVIC date fields empty.")

    # Execute system record lookup (Pure data)
    ok, sql_cols, sql_rows, msg = run_filtered_query(
        server=server,
        database=database,
        username=username,
        password=password,
        auth_type=auth_type,
        driver=driver,
        start_date=start_date,
        end_date=end_date,
        log_callback=log_callback,
    )
    if not ok:
        return False, None, 0, 0, f"Record retrieval failed: {msg}"

    # Build lookup dictionary: clean(RegNo) -> dict of fields
    try:
        reg_idx = sql_cols.index("RegNo")
        debit_idx = sql_cols.index("DebitNote")
        issue_idx = sql_cols.index("Issue Date") if "Issue Date" in sql_cols else None
        pure_start_idx = sql_cols.index("StartDate") if "StartDate" in sql_cols else None
        pure_end_idx = sql_cols.index("EndDate") if "EndDate" in sql_cols else None
    except ValueError as err:
        log_callback(f"ERROR: Missing required fields in system output: {err}")
        return False, None, 0, 0, str(err)


    # Build lookup dictionary: clean(RegNo) -> dict of fields
    pure_lookup: dict[str, dict] = {}
    for r in sql_rows:
        reg_key = clean_reg_no(r[reg_idx])
        if not reg_key:
            continue
        # Extract needed fields from SQL result row
        debit_note = r[debit_idx] if debit_idx is not None else ""
        pure_start = r[pure_start_idx] if pure_start_idx is not None else ""
        pure_end = r[pure_end_idx] if pure_end_idx is not None else ""
        period = f"{pure_start} - {pure_end}" if pure_start and pure_end else ""
        pure_lookup[reg_key] = {
            "PureRegNo": reg_key,
            "DebitNote": debit_note,
            "PureStartDate": pure_start,
            "PureEndDate": pure_end,
            "DebitNotePeriod": period,
        }

    # Prepare new output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "CombinedReport"
    headers = [
        "Pure RegNo",
        "DMVIC RegNo",
        "Debit Note",
        "DMVIC Date",
        "Pure StartDate",
        "Pure EndDate",
        "DebitNote Period",
        "Match Found",
    ]
    out_ws.append(headers)

    matched_count = 0
    total_data_rows = 0
    # Iterate DMVIC rows (starting after header_row_found)
    for row in range(header_row_found + 1, ws.max_row + 1):
        total_data_rows += 1
        dmvic_reg_raw = ws.cell(row=row, column=reg_col_idx).value
        dmvic_reg = clean_reg_no(dmvic_reg_raw)
        dmvic_date = ws.cell(row=row, column=date_col_idx).value if date_col_idx else ""
        lookup = pure_lookup.get(dmvic_reg)
        if lookup:
            matched_count += 1
            out_ws.append([
                lookup.get("PureRegNo", ""),
                dmvic_reg_raw,
                lookup.get("DebitNote", ""),
                dmvic_date,
                lookup.get("PureStartDate", ""),
                lookup.get("PureEndDate", ""),
                lookup.get("DebitNotePeriod", ""),
                "Yes",
            ])
        else:
            out_ws.append([
                "",
                dmvic_reg_raw,
                "",
                dmvic_date,
                "",
                "",
                "",
                "No",
            ])

    log_callback(f"Extraction complete: Matched {matched_count} out of {total_data_rows} DMVIC rows.")

    # Save output file
    try:
        log_callback(f"Saving combined report to: {output_excel_path}")
        out_wb.save(output_excel_path)
        log_callback(f"SUCCESS: Saved combined report to {output_excel_path}")
    except Exception as save_err:
        log_callback(f"ERROR saving combined report: {save_err}")
        return False, None, matched_count, total_data_rows, str(save_err)

    # Convert to pandas DataFrame for UI grid preview (read the newly created sheet)
    df_preview = None
    if pd is not None:
        try:
            df_preview = pd.read_excel(output_excel_path, sheet_name="CombinedReport")
        except Exception:
            pass

    return True, df_preview, matched_count, total_data_rows, f"Combined report generated with {matched_count}/{total_data_rows} matches."

    log_callback(f"Extraction complete: Matched {matched_count} out of {total_data_rows} rows. Written to Column T starting at T5.")

    # Save output file
    try:
        log_callback(f"Saving updated report to: {output_excel_path}")
        wb.save(output_excel_path)
        log_callback(f"SUCCESS: Saved updated report to {output_excel_path}")
    except Exception as save_err:
        log_callback(f"ERROR saving report file: {save_err}")
        return False, None, matched_count, total_data_rows, str(save_err)

    # Convert to pandas DataFrame for UI grid preview
    df_preview = None
    if pd is not None:
        try:
            df_preview = pd.read_excel(output_excel_path, header=3)  # Row 4 is header (0-indexed 3)
        except Exception:
            pass

    return True, df_preview, matched_count, total_data_rows, f"Successfully set DEBITNOTE on T4 and populated data rows ({matched_count}/{total_data_rows} matched)."
